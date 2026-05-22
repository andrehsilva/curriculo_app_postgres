import os
import re
import io
import json
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
from functools import wraps
from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dev-secret-key")

# Nova string de conexão para o Postgres nativo na VPS
DATABASE_URL = os.getenv("DATABASE_URL")

if not DATABASE_URL:
    raise ValueError(
        "A variável de ambiente DATABASE_URL é obrigatória. "
        "Verifique o arquivo .env"
    )

def obter_conexao():
    """Abre uma conexão limpa com o banco de dados PostgreSQL na VPS."""
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)

SENHA_ACESSO_SISTEMA = "piqueaprende2026"
SENHA_EXCLUSAO = "piqueaprende2026"

PROFISSOES = [
    "Professor",
    "Auxiliar",
    "Coordenador",
    "Monitor",
    "Psicopedagogo",
    "Nutricionista",
    "Administrativo",
    "Secretário",
    "Porteiro",
    "Cozinheiro",
]

def login_requerido(f):
    """Decorador para proteger rotas. Redireciona para o login se não estiver autenticado."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("autenticado"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        senha_digitada = request.form.get("senha", "")
        if senha_digitada == SENHA_ACESSO_SISTEMA:
            session["autenticado"] = True
            session.permanent = True  
            return redirect(url_for("index"))
        else:
            return render_template("login.html", erro="Senha incorreta. Tente novamente.")
            
    return render_template("login.html", erro=None)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─────────────────────────────────────────
#  Helpers de Formatação e Tratamento
# ─────────────────────────────────────────

def _fone_para_whatsapp(telefone: str) -> str:
    if not telefone:
        return ""
    digitos = re.sub(r"\D", "", telefone)
    if not digitos.startswith("55"):
        digitos = "55" + digitos
    return f"https://wa.me/{digitos}"


def _normalizar_habilidades(habilidades) -> list:
    if not habilidades:
        return []

    if isinstance(habilidades, str):
        try:
            parsed = json.loads(habilidades)
            if isinstance(parsed, list):
                return [str(h).strip() for h in parsed if h]
        except (json.JSONDecodeError, ValueError):
            pass
        return [h.strip() for h in habilidades.split(",") if h.strip()]

    if isinstance(habilidades, list):
        result = []
        for item in habilidades:
            if isinstance(item, str):
                stripped = item.strip()
                if stripped.startswith("["):
                    try:
                        nested = json.loads(stripped)
                        if isinstance(nested, list):
                            result.extend([str(h).strip() for h in nested if h])
                            continue
                    except (json.JSONDecodeError, ValueError):
                        pass
                result.append(stripped)
            elif item:
                result.append(str(item).strip())
        return [h for h in result if h]

    return []


def _enriquecer_candidato(c: dict) -> dict:
    """Normaliza campos vindos do banco convertendo registros brutos para dicionários amigáveis."""
    c_dit = dict(c)
    c_dit["_habilidades"] = _normalizar_habilidades(c_dit.get("principais_habilidades"))
    c_dit["_whatsapp_url"] = _fone_para_whatsapp(c_dit.get("telefone", ""))
    
    # Tratamento para impedir o erro 'datetime' object is not subscriptable no HTML
    data_raw = c_dit.get("created_at")
    if isinstance(data_raw, datetime):
        c_dit["created_at"] = data_raw.strftime('%Y-%m-%d %H:%M:%S')
    elif data_raw:
        c_dit["created_at"] = str(data_raw)
    else:
        c_dit["created_at"] = ""
        
    return c_dit


def _coletar_todas_tags(candidatos: list) -> list:
    tags = set()
    for c in candidatos:
        for h in c.get("_habilidades", []):
            if h:
                tags.add(h)
    return sorted(tags)


# ─────────────────────────────────────────
#  Rotas Principais
# ─────────────────────────────────────────

@app.route("/")
@login_requerido
def index():
    filtro      = request.args.get("filtro", "todos")
    selecionado = request.args.get("selecionado", None)
    profissao   = request.args.get("profissao", "")
    tag         = request.args.get("tag", "")

    todos_candidatos = _buscar_candidatos(filtro, profissao)
    todas_tags = _coletar_todas_tags(todos_candidatos)

    candidatos = _buscar_candidatos(filtro, profissao, tag)

    candidato_ativo = None
    if candidatos:
        if selecionado:
            candidato_ativo = next(
                (c for c in candidatos if str(c.get("id")) == str(selecionado)),
                candidatos[0],
            )
        else:
            candidato_ativo = candidatos[0]

    return render_template(
        "index.html",
        candidatos=candidatos,
        candidato_ativo=candidato_ativo,
        filtro=filtro,
        profissao=profissao,
        profissoes=PROFISSOES,
        tag=tag,
        todas_tags=todas_tags,
    )


@app.route("/partial/sidebar")
@login_requerido
def partial_sidebar():
    filtro      = request.args.get("filtro", "todos")
    selecionado = request.args.get("selecionado", None)
    profissao   = request.args.get("profissao", "")
    tag         = request.args.get("tag", "")

    candidatos = _buscar_candidatos(filtro, profissao, tag)
    candidato_ativo_id = selecionado or (str(candidatos[0]["id"]) if candidatos else None)

    return render_template(
        "partials/sidebar_cards.html",
        candidatos=candidatos,
        candidato_ativo_id=candidato_ativo_id,
        filtro=filtro,
        profissao=profissao,
        tag=tag,
    )


@app.route("/api/curriculos")
def api_curriculos():
    filtro    = request.args.get("filtro", "todos")
    profissao = request.args.get("profissao", "")
    tag       = request.args.get("tag", "")

    candidatos = _buscar_candidatos(filtro, profissao, tag)

    return jsonify({
        "total": len(candidatos),
        "ids":   [str(c["id"]) for c in candidatos],
    })


def _buscar_candidatos(filtro: str, profissao: str, tag: str = "") -> list:
    """Busca candidatos no Postgres aplicando os filtros e ordena pela maior compatibilidade (%)."""
    query = "SELECT * FROM public.triagem_curriculos WHERE 1=1"
    params = []

    if filtro == "aprovados":
        query += " AND aprovado = TRUE"
    elif filtro == "reprovados":
        query += " AND aprovado = FALSE"

    if profissao and profissao.strip():
        query += " AND cargo_alvo ILIKE %s"
        params.append(f"%{profissao.strip()}%")

    # 1. IMPORTANTE: Inicializar como None ANTES do try
    conn = None
    cursor = None
    linhas = []

    try:
        conn = obter_conexao()
        cursor = conn.cursor()
        cursor.execute(query, params)
        linhas = cursor.fetchall()
    except Exception as e:
        # 2. Isso vai cuspir no terminal o real motivo de não estar conectando
        print(f"[Postgres] Erro crítico de conexão/query: {e}")
        return []
    finally:
        # 3. Garante o fechamento seguro sem gerar UnboundLocalError
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    # Enriquecimento dos metadados locais (habilidades e whatsapp)
    candidatos = [_enriquecer_candidato(row) for row in linhas]

    if tag:
        tag_lower = tag.lower()
        candidatos = [
            c for c in candidatos
            if any(tag_lower in h.lower() for h in c.get("_habilidades", []))
        ]

    def obter_nota(candidato):
        try:
            return float(candidato.get("nota_compatibilidade", 0))
        except (ValueError, TypeError):
            return 0.0

    candidatos.sort(key=obter_nota, reverse=True)
    return candidatos


@app.route("/api/contratar/<candidato_id>", methods=["POST"])
def contratar_candidato(candidato_id):
    """Muda o status do candidato para Aprovado (aprovado = True) usando SQL Nativo."""
    try:
        conn = obter_conexao()
        cursor = conn.cursor()
        
        query = "UPDATE triagem_curriculos SET aprovado = TRUE WHERE id = %s"
        cursor.execute(query, (candidato_id,))
        
        # Confirma as alterações na transação do Postgres
        conn.commit()
        linhas_afetadas = cursor.rowcount
        
        if linhas_afetadas == 0:
            return jsonify({"ok": False, "message": "Candidato não encontrado ou nenhuma alteração feita."}), 404

        return jsonify({"ok": True, "message": "Candidato aprovado com sucesso!"})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route("/api/analisar/<candidato_id>", methods=["POST"])
def analisar_candidato(candidato_id):
    """Muda o status do candidato de volta para Em Análise (aprovado = False) usando SQL Nativo."""
    try:
        conn = obter_conexao()
        cursor = conn.cursor()
        
        query = "UPDATE triagem_curriculos SET aprovado = FALSE WHERE id = %s"
        cursor.execute(query, (candidato_id,))
        
        conn.commit()
        linhas_afetadas = cursor.rowcount
        
        if linhas_afetadas == 0:
            return jsonify({"ok": False, "message": "Candidato não encontrado ou nenhuma alteração feita."}), 404

        return jsonify({"ok": True, "message": "Candidato movido para Em Análise!"})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route("/api/excluir/<candidato_id>", methods=["POST"])
def excluir_candidato(candidato_id):
    """Exclui candidato permanentemente via query DELETE pós verificação de senha."""
    dados = request.get_json(silent=True) or {}
    senha = dados.get("senha", "")

    if senha != SENHA_EXCLUSAO:
        return jsonify({"ok": False, "message": "Senha incorreta."}), 403

    try:
        conn = obter_conexao()
        cursor = conn.cursor()
        
        query = "DELETE FROM triagem_curriculos WHERE id = %s"
        cursor.execute(query, (candidato_id,))
        
        conn.commit()
        return jsonify({"ok": True, "message": "Candidato excluído com sucesso!"})
    except Exception as e:
        print(f"[Postgres] Erro ao excluir registro: {e}")
        return jsonify({"ok": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


# ─────────────────────────────────────────
#  Exportar para Excel (Independente de ORM)
# ─────────────────────────────────────────

@app.route("/exportar")
@login_requerido
def exportar_excel():
    candidatos = _buscar_candidatos(filtro="todos", profissao="", tag="")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Todos os Candidatos"

    cor_header   = "B85C72"  
    cor_aprovado = "E8F8EF"  
    cor_analise  = "FEF3C7"  

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor=cor_header)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="EBEBEB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    colunas = [
        ("Nome",                 25),
        ("Idade",                8),
        ("Sexo",                 10),
        ("Email",                28),
        ("Telefone",             16),
        ("Endereço",             30),
        ("Cargo Alvo",           20),
        ("Compatibilidade (%)",  18),
        ("Status (Aprovado)",    16),
        ("Classificação",        22),
        ("Habilidades",          45),
        ("Resumo Profissional",  50),
        ("Alertas de Risco",     30),
        ("Data de Recebimento",  20),
    ]

    ws.row_dimensions[1].height = 30
    for col_idx, (label, width) in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, c in enumerate(candidatos, start=2):
        habilidades = ", ".join(c.get("_habilidades", []))
        status_texto = "Aprovado" if c.get("aprovado") else "Em Análise"
        
        # Garante tratamento limpo se a data vier como objeto datetime ou string nativa
        data_raw = c.get("created_at", "")
        if isinstance(data_raw, datetime):
            data_fmt = data_raw.strftime('%Y-%m-%d')
        else:
            data_fmt = str(data_raw)[:10] if data_raw else ""

        valores = [
            c.get("nome_candidato", ""),
            c.get("idade", ""),
            c.get("sexo", ""),
            c.get("email", ""),
            c.get("telefone", ""),
            c.get("endereco", ""),
            c.get("cargo_alvo", ""),
            c.get("nota_compatibilidade", ""),
            status_texto,
            c.get("classificacao", ""),
            habilidades,
            c.get("resumo_professional", "") or c.get("resumo_profissional", ""),
            c.get("alertas_de_risco", ""),
            data_fmt,
        ]

        fill_color = cor_aprovado if c.get("aprovado") else cor_analise
        row_fill = PatternFill("solid", fgColor=fill_color)

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.row_dimensions[row_idx].height = 40

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"total_candidatos_pique_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nome_arquivo,
    )

if __name__ == "__main__":
    debug = os.getenv("FLASK_ENV", "production") == "development"
    app.run(debug=debug, host="0.0.0.0", port=5000)